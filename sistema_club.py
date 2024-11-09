from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
import os

class FichaMedica:
    def __init__(self):
        self.ficha = None
        
    def abrir_ficha(self):
        # Si ya hay una ventana abierta, no abrir otra
        if self.ficha is not None and self.ficha.winfo_exists():
            self.ficha.lift()  # Trae la ventana al frente
            return
            
        # Inicializa la ventana de Tkinter
        self.ficha = Toplevel()
        self.ficha.geometry("500x600")
        self.ficha.configure(bg="dark blue")
        self.ficha.title("Ficha Médica")
        
        # Configura la ruta del archivo
        self.workbook_path = "h:/2024/abi/EXCEL(BASE DE DATOS.T).xlsx"
        
        # Verifica el archivo antes de continuar
        if not self.verificar_archivo():
            messagebox.showwarning("Advertencia", "Usando modo de prueba. Los datos no se guardarán.")
        
        try:
            self.workbook = load_workbook(self.workbook_path)
            self.sheet = self.workbook["ficha medica"]
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar el archivo Excel: {str(e)}")
            self.ficha.destroy()
            return
            
        self.crear_widgets()
    
    def verificar_archivo(self):
        try:
            if not os.path.exists(self.workbook_path):
                wb = load_workbook()
                wb.create_sheet("ficha medica")
                wb.save(self.workbook_path)
                return True
                
            with open(self.workbook_path, 'rb+') as f:
                return True
        except PermissionError:
            messagebox.showerror("Error", "No hay permisos para acceder al archivo Excel.")
            return False
        except Exception as e:
            messagebox.showerror("Error", f"Error al acceder al archivo: {str(e)}")
            return False
    
    def descomprimir_celdas(self):
        try:
            merged_ranges = list(self.sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                self.sheet.unmerge_cells(str(merged_range))
        except Exception as e:
            print(f"Error al descomprimir celdas: {str(e)}")
    
    def guardar_datos(self):
        if not self.verificar_archivo():
            return
            
        try:
            # Obtiene el índice de la celda A3 y lo incrementa
            index = self.sheet["A3"].value
            if index is None:
                index = 1
            else:
                index += 1
            
            self.descomprimir_celdas()
            
            # Actualiza el índice
            self.sheet["A3"].value = index
            
            # Guarda los datos
            self.sheet[f"B{index}"] = self.entry_nombre.get()
            self.sheet[f"C{index}"] = self.entry_apellido.get()
            self.sheet[f"D{index}"] = self.entry_dni.get()
            self.sheet[f"E{index}"] = self.entry_alergias.get()
            self.sheet[f"F{index}"] = self.entry_cirugias.get()
            self.sheet[f"G{index}"] = self.entry_lesiones.get()
            
            self.workbook.save(self.workbook_path)
            messagebox.showinfo("Éxito", "Datos guardados correctamente")
            
            # Limpia los campos
            for entry in [self.entry_nombre, self.entry_apellido, self.entry_dni, 
                         self.entry_alergias, self.entry_cirugias, self.entry_lesiones]:
                entry.delete(0, END)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar los datos: {str(e)}")
    
    def crear_widgets(self):
        # Etiquetas y campos de entrada
        Label(self.ficha, text="Información Deportiva", bg="dark blue", fg="white", 
              font=("Lato", 13)).place(x=180, y=50)
        Button(self.ficha, text="Guardar", command=self.guardar_datos, 
               bg="dark blue", fg="white", font=("Lato", 13)).place(x=180, y=350)

        # Crear y ubicar las etiquetas y campos
        campos = [
            ("Nombre:", "entry_nombre"),
            ("Apellido:", "entry_apellido"),
            ("DNI:", "entry_dni"),
            ("Alergias:", "entry_alergias"),
            ("Cirugías Previas:", "entry_cirugias"),
            ("Lesiones Previas:", "entry_lesiones")
        ]
        
        for i, (label_text, entry_name) in enumerate(campos):
            Label(self.ficha, text=label_text, bg="dark blue", fg="white", 
                  font=("Lato", 12)).place(x=10, y=100 + i*40)
            entry = Entry(self.ficha)
            entry.place(x=150, y=100 + i*40)
            setattr(self, entry_name, entry)

class AplicacionPrincipal:
    def __init__(self):
        self.ventana = Tk()
        self.ventana.geometry("500x600")
        self.ventana.configure(bg="dark blue")
        self.ventana.title("Sistema del Club")
        
        # Instancia de FichaMedica
        self.ficha_medica = FichaMedica()
        
        self.crear_menu()
        
    def crear_menu(self):
        menu_superior = Menu(self.ventana)
        self.ventana.config(menu=menu_superior)
        
        # Crear menús
        jugadores = Menu(menu_superior, tearoff=0)
        socios = Menu(menu_superior, tearoff=0)
        ayuda = Menu(menu_superior, tearoff=0)
        salir = Menu(menu_superior, tearoff=0)
        Categoria = Menu(menu_superior, tearoff=0)
        Ninos = Menu(Categoria, tearoff=0)
        
        # Agregar las opciones principales
        menu_superior.add_cascade(label="Jugador", menu=jugadores)
        menu_superior.add_cascade(label="Funcionalidad", menu=Categoria)
        menu_superior.add_cascade(label="Socio", menu=socios)
        menu_superior.add_cascade(label="Ayuda", menu=ayuda)
        menu_superior.add_cascade(label="Salir", menu=salir)
        
        # Configurar submenús
        jugadores.add_command(label="ficha medica", command=self.ficha_medica.abrir_ficha)
        jugadores.add_separator()
        jugadores.add_command(label="legajo del jugador")
        
        socios.add_command(label="Beneficios")
        socios.add_separator()
        socios.add_command(label="pagos")
        
        ayuda.add_command(label="Acerca de...")
        ayuda.add_separator()
        ayuda.add_command(label="Historia del club")
        
        Categoria.add_cascade(label="Informacion deportiva", menu=Ninos)
    
    def iniciar(self):
        self.ventana.mainloop()

if __name__ == "__main__":
    app = AplicacionPrincipal()
    app.iniciar()